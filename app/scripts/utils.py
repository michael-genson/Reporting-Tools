from tempfile import NamedTemporaryFile

import openpyxl
import pandas as pd
from openpyxl import Workbook
from streamlit.uploaded_file_manager import UploadedFile


def format_percent(val: float | int) -> str:
    return f"{round(val * 100, 2)}%"


def open_workbook(report_file: UploadedFile, file_display_name: str) -> Workbook:
    # XLSX
    try:
        report_file.seek(0)
        return openpyxl.load_workbook(report_file)

    except Exception:
        pass

    # HTML
    try:
        report_file.seek(0)
        df = pd.read_html(report_file)[0]

        with NamedTemporaryFile(delete=False) as f:
            df.to_excel(f, engine="openpyxl")
            return openpyxl.load_workbook(f)  # type: ignore

    except Exception:
        pass

    # XLS
    try:
        report_file.seek(0)
        df = pd.read_excel(report_file, engine="xlrd")

        with NamedTemporaryFile(delete=False) as f:
            df.to_excel(f, engine="openpyxl")
            return openpyxl.load_workbook(f)  # type: ignore

    except Exception:
        pass

    # CSV
    try:
        report_file.seek(0)
        df = pd.read_csv(report_file)

        with NamedTemporaryFile(delete=False) as f:
            df.to_excel(f, engine="openpyxl")
            return openpyxl.load_workbook(f)  # type: ignore

    except Exception:
        pass

    raise ValueError(
        f"Unable to open {file_display_name} file; is it in the right format?"
    )
