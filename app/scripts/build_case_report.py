from datetime import datetime
from tempfile import NamedTemporaryFile

from openpyxl.styles import Alignment
from openpyxl.styles.borders import Border, Side
from openpyxl.styles.fills import PatternFill
from openpyxl.worksheet.worksheet import Worksheet
from static import follow_up_and_weight_map

from scripts.utils import open_workbook


def load_follow_up_and_weight_map() -> dict[str, int | float]:
    d = follow_up_and_weight_map.get_map()

    # Set all keys to lowercase
    d = {k.strip().lower(): v for k, v in d.items()}
    return d


def write_workbook_runtime(ws: Worksheet, report_datetime: datetime) -> None:
    """Appends the runtime to the workbook"""

    datetime_string = report_datetime.strftime("%-m/%-d/%Y at %-I:%M%p").lower()

    for i, row in enumerate(ws.rows):
        if row[0].value and "workload management report" in row[0].value.lower():
            row[0].value += datetime_string

        elif i > 100:
            raise ValueError(
                "Cannot find cell to write runtime in. Is the file formatted correctly?"
            )


def parse_pivot_table_for_cycles_and_weights(
    ws: Worksheet, follow_up_weight_map: dict
) -> dict:
    """
    Parses the pivot table data to write cycles and weights and blackens unused cells

    Returns a map of agent name to row count and average cycles
    """

    row_labels_index = 0
    avg_case_age_index = 1

    cycles_index = 3
    weights_index = 4

    cell_alignment = Alignment(horizontal="center")

    last_agent = None
    agent_row_count = 0
    agent_cycle_count = 0
    agent_data_map = {}

    start_parsing = False
    for row in ws.rows:
        # don't start parsing until we get to the pivot table
        if not start_parsing:
            if row[row_labels_index].pivotButton:
                start_parsing = True

            continue

        # stop when we've reached the end of the pivot table
        if row[row_labels_index].value == "Grand Total":
            break

        # don't run any calculations when we've reached a new agent
        if row[row_labels_index].alignment.indent.real == 0.0:
            # map agent and reset agent vars
            if last_agent:
                agent_data_map[last_agent] = {
                    "row_count": agent_row_count,
                    "average_cycles": float(
                        round(agent_cycle_count / agent_row_count, 2)
                    ),
                }
            last_agent = row[row_labels_index].value
            agent_row_count = 0
            agent_cycle_count = 0

            # black-out unused cycle and weight cells
            for index in [cycles_index, weights_index]:
                row[index].fill = PatternFill(
                    start_color="00000000", end_color="00000000", fill_type="solid"
                )

            continue

        # write cycles and weights
        label = row[row_labels_index].value.strip().lower()

        cycle_cell = row[cycles_index]
        cycle_cell.value = f"={row[avg_case_age_index].coordinate}/{follow_up_weight_map[label]['follow-up']}"
        cycle_cell.number_format = "0.00"
        cycle_cell.alignment = cell_alignment

        weight_cell = row[weights_index]
        weight_cell.value = (
            f"={row[cycles_index].coordinate}*{follow_up_weight_map[label]['weight']}"
        )
        weight_cell.number_format = "0.00"
        weight_cell.alignment = cell_alignment

        # track agent rows and cycle count using the weight column calculation
        cycle_cell_calculation = (
            row[avg_case_age_index].value / follow_up_weight_map[label]["follow-up"]
        )
        weight_cell_calculation = (
            cycle_cell_calculation * follow_up_weight_map[label]["weight"]
        )

        agent_row_count += 1
        agent_cycle_count += weight_cell_calculation

    # map final agent and return map
    agent_data_map[last_agent] = {
        "row_count": agent_row_count,
        "average_cycles": float(round(agent_cycle_count / agent_row_count, 2)),
    }

    return agent_data_map


def write_average_cycle_formulas(
    ws: Worksheet, agent_data_map: dict, weight_performance_map: dict
):
    """Writes the average weight formulas and paints them per the weight performance map"""

    row_labels_index = 0
    weights_index = 4
    avg_cycles_index = 5

    cell_alignment = Alignment(horizontal="center")

    start_parsing = False
    for i, row in enumerate(ws.rows):
        # don't start parsing until we get to the pivot table
        if not start_parsing:
            if row[row_labels_index].pivotButton:
                start_parsing = True

            continue

        # stop when we've reached the end of the pivot table
        if row[row_labels_index].value == "Grand Total":
            break

        # ignore rows that aren't a new agent
        if row[row_labels_index].alignment.indent.real != 0.0:
            continue

        # get agent and number of rows to average
        agent = row[row_labels_index].value
        row_count = agent_data_map[agent]["row_count"]

        # write average formula
        avg_cycle_cell = ws[i + 2][avg_cycles_index]

        start_cell = ws[i + 2][weights_index].coordinate
        end_cell = ws[i + row_count + 1][weights_index].coordinate

        avg_cycle_cell.value = f"=AVERAGE({start_cell}:{end_cell})"
        avg_cycle_cell.number_format = "0.00"

        # format cell based on weight performance map
        avg_cycle_cell.alignment = cell_alignment

        for max_threshold, color in weight_performance_map.items():
            if (
                max_threshold == "default"
                or agent_data_map[agent]["average_cycles"] < max_threshold
            ):
                cell_color = f"FF{color[-6:]}"  # convert to aRGB
                break

        avg_cycle_cell.fill = PatternFill(
            start_color=cell_color, end_color=cell_color, fill_type="solid"
        )

        border_side = Side(style="medium")
        avg_cycle_cell.border = Border(
            left=border_side, right=border_side, top=border_side, bottom=border_side
        )


def main(
    report_file,
    report_datetime: datetime,
    outstanding_val: float,
    outstanding_color: str,
    exceeds_val: float,
    exceeds_color: str,
    competent_val: float,
    competent_color: str,
    needs_improvement_color: str,
):

    follow_up_weight_map = load_follow_up_and_weight_map()

    weight_performance_map: dict[float | str, str] = {
        float(outstanding_val): outstanding_color,
        float(exceeds_val): exceeds_color,
        float(competent_val): competent_color,
        "default": needs_improvement_color,
    }

    wb = open_workbook(report_file, "Case Report")
    ws = wb.active
    write_workbook_runtime(ws, report_datetime)

    try:
        agent_data_map = parse_pivot_table_for_cycles_and_weights(
            ws, follow_up_weight_map
        )

        write_average_cycle_formulas(ws, agent_data_map, weight_performance_map)

    except Exception as e:
        print(e)
        raise ValueError(
            "Unable to format workbook. Is the file formatted correctly?"
        ) from e

    f = NamedTemporaryFile(delete=False)
    wb.save(f.name)

    return f.name
